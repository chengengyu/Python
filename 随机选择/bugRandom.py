__author__ = 'chengengyu'
from openpyxl import Workbook, load_workbook
import copy
import random

class BugInfo(object):
    def __init__(self, BugNum, OpenInfo, ResolveInfo, CloseInfo):
        self.BugNum = BugNum
        self.OpenInfo = OpenInfo
        self.ResolveInfo = ResolveInfo
        self.CloseInfo = CloseInfo


SourceFileName = input("本次抽查xls: ")
BugNum = int(input("本次抽查个数："))
#读取源表
SourWb = load_workbook(SourceFileName)
SourWs = SourWb.active

#RandomList = random.sample(SourWs.rows, BugNum)
#DesWb = Workbook()
#DesWs = DesWb.active
#DesWs.title = "原始数据"
DesWs = SourWb.create_sheet("抽查结果")
#将抽查的内容拷贝到一个新的表中
#df = DataFrame(ws.values)
rowList = []
cellList = []
for rowNum, row in enumerate(SourWs.rows):
    cellList = []
    for cell in row:
        cellList.append(cell.value)
    rowList.append(cellList)

RandomList = random.sample(rowList, BugNum)

for rowNum, row in enumerate(RandomList):
    for cellNum, cell in enumerate(row):
        DesWs.cell(row=rowNum+1, column=cellNum+1).value = cell
SourWb.save(SourceFileName)

input("All Done, press any key to continue.")

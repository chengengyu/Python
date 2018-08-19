from splinter import Browser
import datetime
from openpyxl import Workbook, load_workbook

#htmlDir = r'K:\\work\\GitHub\\Python\\考勤抓取\\大唐移动考勤请假管理系统.html'

#html页面对应信息的ID
idList = [\
    'lbshenqingren',\
    'lbshenqingshijian',\
    'lbjiaqishijian',\
    'lbkuadu',\
    'lbjiabie',\
    'lbshiyou',\
    'lbshenpizhuangtai',\
    'lbshenpilishi',\
    'lbbeizhu'\
]

#写入excel的时候标题，比idList多了第一位'序号'
titleList = [\
    '序号',\
    '申请人',\
    '申请时间',\
    '假期时间',\
    '跨度',\
    '假别',\
    '事由',\
    '审批状态',\
    '审批历史',\
    '备注'\
]

#从第一页面都到最后一页，获取所有页面假单页面的链接
def getHref (browser):
    hrefList = []
    elementList = browser.find_link_by_partial_text('查看')
    for ele in elementList:
        #print(ele['href'])
        hrefList.append(ele['href'])
    nextHref = b.find_link_by_partial_text('下一页')
    if nextHref:
        nextHref[0].click()
        hrefList = hrefList + getHref(browser)
    return hrefList

#根据配置的ID，获取假单页面的信息
def getInfo(browser):
    infoDic = {}
    for eleId in idList:
        ele = browser.find_by_id(eleId)
        infoDic[eleId] = ele.value
    return infoDic

#将获取的到的所有假单写入excel
def writeExcel(infoList):
    SaveWb = Workbook()
    SaveWs = SaveWb.active
    SaveWs.title = '考勤'
    for num, title in enumerate(titleList):
        SaveWs.cell(row=1, column=num + 1, value=title)
    for num, infoDic in enumerate(infoList):
        SaveWs.cell(row=num+2, column=1, value=num+1)
        for idNum, eleId in enumerate(idList):
            SaveWs.cell(row=num+2, column=idNum+2, value=infoDic[eleId])
    SaveWb.save('考勤假单.xlsx')
    

if __name__ =="__main__":
    htmlDir = input("请输入网址：")
    url = htmlDir
    b = Browser(driver_name='chrome')
    b.visit(url)
    hrefList = getHref(b)
    print(len(hrefList))
    infoList = []
    for href in hrefList:
        try:
            b.visit(href)
            infoList.append(getInfo(b))
        except:
            print('无法访问：' + href)
    print(infoList)    
    writeExcel(infoList)



